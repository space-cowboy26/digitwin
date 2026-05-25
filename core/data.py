# core/data.py

import logging
from pathlib import Path
from typing import List

import traceback

import numpy as np
import pandas as pd
from openpyxl import load_workbook

from config.settings import (
    GII_FILTER, TIME_START, TIME_END,
    GAP_ROWS, VAL_DAYS, TEST_DAYS,
    EXCLUDE_FROM_FEATURES,
    SHEET_NAME_MAP,
    ROLLING_WINDOW_TRIGGER_MONTHS,
    ROLLING_WINDOW_DAYS,
)

log = logging.getLogger(__name__)


# -- Header detection ----------------------------------------------------------

def _find_header_row(ws) -> int:
    """
    Scan worksheet rows until a cell containing 'time' or 'stamp' is found.
    Returns 0-indexed row number of the header row.
    """
    for i, row in enumerate(ws.iter_rows(max_row=25, values_only=True)):
        for cell in row:
            if cell and isinstance(cell, str):
                if "time" in cell.lower() or "stamp" in cell.lower():
                    return i
    raise ValueError(
        "Could not detect header row. "
        "Expected a column containing 'time' or 'stamp' in first 25 rows."
    )


# -- Individual file readers --------------------------------------------------─

def _read_inverter_file(filepath, itc_inv: str) -> pd.DataFrame:
    """
    Read one daily inverter report xlsx.
    Detects header row automatically.
    Reads the sheet corresponding to itc_inv.
    Drops units row (always immediately after header).
    """
    wb         = load_workbook(filepath, read_only=True, data_only=True)
    sheet_name = SHEET_NAME_MAP.get(itc_inv)

    if sheet_name not in wb.sheetnames:
        raise ValueError(
            f"Sheet '{sheet_name}' not found in {Path(filepath).name}. "
            f"Available sheets: {wb.sheetnames}"
        )

    ws         = wb[sheet_name]
    header_row = _find_header_row(ws)
    wb.close()

    df = pd.read_excel(
        filepath,
        sheet_name = sheet_name,
        skiprows   = header_row,
        engine     = "openpyxl",
    )
    df = df.iloc[1:].reset_index(drop=True)         # drop units row
    df = df.dropna(how="all").reset_index(drop=True)     # drop empty rows
    df = df.dropna(how="all", axis=1).reset_index(drop=True)  # drop empty cols

    ts_col = _detect_timestamp_col(df)
    df, ts_warning = _parse_file_timestamps(df, ts_col)
    if ts_warning:
        log.warning(f"Inverter file {Path(str(filepath)).name}: {ts_warning}")

    return df, ts_warning


def _read_wms_file(filepath) -> pd.DataFrame:
    """
    Read one daily WMS report xlsx.
    Detects header row automatically.
    Drops units row (always immediately after header).
    """
    wb         = load_workbook(filepath, read_only=True, data_only=True)
    ws         = wb.active
    header_row = _find_header_row(ws)
    wb.close()

    df = pd.read_excel(
        filepath,
        sheet_name = 0,
        skiprows   = header_row,
        engine     = "openpyxl",
    )
    ddf = df.iloc[1:].reset_index(drop=True)
    df = df.dropna(how="all").reset_index(drop=True)
    df = df.dropna(how="all", axis=1).reset_index(drop=True)

    ts_col = _detect_timestamp_col(df)
    df, ts_warning = _parse_file_timestamps(df, ts_col)
    if ts_warning:
        log.warning(f"WMS file {Path(str(filepath)).name}: {ts_warning}")

    return df, ts_warning


# -- Multi-file loader and merger ----------------------------------------------

def load_and_merge(
    inv_filepaths: List,
    wms_filepaths: List,
    itc_inv:       str,
    fill_gaps:     bool = True,
) -> tuple:
    """
    1. Read each inverter file -> extract correct sheet -> detect headers
    2. Read each WMS file -> detect headers
    3. Concatenate all inverter frames, all WMS frames separately
    4. Deduplicate and sort both by timestamp
    5. Outer merge inverter + WMS on timestamp
    6. Fill small gaps (<=5 min) for inference, drop missing for train/retrain
    7. Return (merged DataFrame, list of warnings)
    """
    log.info(
        f"Loading {len(inv_filepaths)} inverter files "
        f"and {len(wms_filepaths)} WMS files for {itc_inv}"
    )
    merge_warnings = []

    # flatten in case of nested lists
    if inv_filepaths and isinstance(inv_filepaths[0], list):
        inv_filepaths = [fp for sublist in inv_filepaths for fp in sublist]
    if wms_filepaths and isinstance(wms_filepaths[0], list):
        wms_filepaths = [fp for sublist in wms_filepaths for fp in sublist]

    inv_filepaths = [Path(fp) for fp in inv_filepaths]
    wms_filepaths = [Path(fp) for fp in wms_filepaths]

    # -- Read inverter files -----------------------------------------------
    inv_frames = []
    for fp in inv_filepaths:
        try:
            df, ts_warn = _read_inverter_file(fp, itc_inv)
            inv_frames.append(df)
            if ts_warn:
                merge_warnings.append(f"{Path(str(fp)).name}: {ts_warn}")
            log.info(f"  Inverter file read: {Path(str(fp)).name} - {len(df)} rows")
        except Exception as e:
            log.warning(f"  Skipping inverter file {Path(str(fp)).name}: {e}\n{traceback.format_exc()}")

    if not inv_frames:
        raise ValueError("No inverter files could be read successfully.")

    # -- Read WMS files ----------------------------------------------------
    wms_frames = []
    for fp in wms_filepaths:
        try:
            df, ts_warn = _read_wms_file(fp)
            wms_frames.append(df)
            if ts_warn:
                merge_warnings.append(f"{Path(str(fp)).name}: {ts_warn}")
            log.info(f"  WMS file read: {Path(str(fp)).name} - {len(df)} rows")
        except Exception as e:
            log.warning(f"  Skipping WMS file {Path(str(fp)).name}: {e}")

    if not wms_frames:
        raise ValueError("No WMS files could be read successfully.")

    # -- Concatenate -------------------------------------------------------
    inv_df = pd.concat(inv_frames, ignore_index=True)
    wms_df = pd.concat(wms_frames, ignore_index=True)

    # -- Detect timestamp columns (pre-normalisation) ----------------------
    inv_ts_col = _detect_timestamp_col(inv_df)
    wms_ts_col = _detect_timestamp_col(wms_df)

    # floor to minute to remove sub-second duplicates
    inv_df[inv_ts_col] = pd.to_datetime(inv_df[inv_ts_col]).dt.floor("min")
    wms_df[wms_ts_col] = pd.to_datetime(wms_df[wms_ts_col]).dt.floor("min")

    # -- Sort and deduplicate ----------------------------------------------
    inv_df = inv_df.sort_values(inv_ts_col).drop_duplicates(
        subset=[inv_ts_col]).reset_index(drop=True)
    wms_df = wms_df.sort_values(wms_ts_col).drop_duplicates(
        subset=[wms_ts_col]).reset_index(drop=True)

    log.info(
        f"After dedup - inverter: {len(inv_df)} rows | "
        f"WMS: {len(wms_df)} rows"
    )

    # -- Rename timestamp cols to common key for merge ---------------------
    inv_df = inv_df.rename(columns={inv_ts_col: "__ts__"})
    wms_df = wms_df.rename(columns={wms_ts_col: "__ts__"})

    # -- Outer merge -------------------------------------------------------
    merged = pd.merge(inv_df, wms_df, on="__ts__", how="outer")
    merged = merged.sort_values("__ts__").reset_index(drop=True)

    # -- Gap handling ------------------------------------------------------
    inv_core_col = "Active Power"
    wms_core_col = "GII"

    if fill_gaps:
        # inference only - fill small gaps, drop large ones
        inv_missing = (
            merged[inv_core_col].isna().sum()
            if inv_core_col in merged.columns else 0
        )
        wms_missing = (
            merged[wms_core_col].isna().sum()
            if wms_core_col in merged.columns else 0
        )

        merged = merged.infer_objects(copy=False).ffill(limit=5).bfill(limit=5)

        inv_filled = inv_missing - (
            merged[inv_core_col].isna().sum()
            if inv_core_col in merged.columns else 0
        )
        wms_filled = wms_missing - (
            merged[wms_core_col].isna().sum()
            if wms_core_col in merged.columns else 0
        )

        if inv_filled > 0:
            msg = (
                f"{inv_filled} inverter rows were missing and filled "
                f"from adjacent timestamps (gap <= 5 minutes)."
            )
            log.warning(msg)
            merge_warnings.append(msg)

        if wms_filled > 0:
            msg = (
                f"{wms_filled} WMS rows were missing and filled "
                f"from adjacent timestamps (gap <= 5 minutes)."
            )
            log.warning(msg)
            merge_warnings.append(msg)

        total_before = len(merged)
        merged = merged.dropna(
            subset=[inv_core_col, wms_core_col]
        ).reset_index(drop=True)
        dropped = total_before - len(merged)
        if dropped > 0:
            msg = f"{dropped} rows dropped - timestamp gaps > 5 consecutive minutes."
            log.warning(msg)
            merge_warnings.append(msg)

    else:
        before = len(merged)
        missing_mask = merged[inv_core_col].isna() | merged[wms_core_col].isna()
        dropped_df   = merged[missing_mask].copy()
        merged       = merged[~missing_mask].reset_index(drop=True)
        dropped      = before - len(merged)

        if dropped > 0:
            # summarise dropped periods
            dropped_ts = pd.to_datetime(dropped_df["__ts__"]).dropna()

            if len(dropped_ts) == 0:
                msg = f"{dropped} rows dropped due to missing or unmatched timestamps."
                log.warning(msg)
                merge_warnings.append(msg)
            else:
                date_range = (
                    f"{dropped_ts.min().strftime('%Y-%m-%d %H:%M')} to "
                    f"{dropped_ts.max().strftime('%Y-%m-%d %H:%M')}"
                )

            # group into contiguous blocks
            dropped_sorted = dropped_ts.sort_values().reset_index(drop=True)
            gaps = []
            block_start = dropped_sorted.iloc[0]
            block_end   = dropped_sorted.iloc[0]
            for i in range(1, len(dropped_sorted)):
                diff = (dropped_sorted.iloc[i] - dropped_sorted.iloc[i-1]).seconds / 60
                if diff <= 2:
                    block_end = dropped_sorted.iloc[i]
                else:
                    gaps.append(f"{block_start.strftime('%Y-%m-%d %H:%M')} to {block_end.strftime('%Y-%m-%d %H:%M')}")
                    block_start = dropped_sorted.iloc[i]
                    block_end   = dropped_sorted.iloc[i]
            gaps.append(f"{block_start.strftime('%Y-%m-%d %H:%M')} to {block_end.strftime('%Y-%m-%d %H:%M')}")

            msg = (
                f"{dropped} rows dropped due to missing or unmatched timestamps. "
                f"Overall range: {date_range}. "
                f"{len(gaps)} gap period(s) detected."
            )
            log.warning(msg)
            merge_warnings.append(msg)

            # append each gap period as a separate warning
            for gap in gaps:
                gap_msg = f"  Gap period: {gap}"
                log.warning(gap_msg)
                merge_warnings.append(gap_msg)
            # -- Rename common key back to original timestamp col name -------------
            merged = merged.rename(columns={"__ts__": inv_ts_col})

            log.info(
                f"Merged DataFrame: {len(merged)} rows, "
                f"{len(merged.columns)} columns"
            )
            return merged, merge_warnings    


def _detect_timestamp_col(df: pd.DataFrame) -> str:
    """
    Find the timestamp column in a raw (pre-normalisation) DataFrame.
    Looks for columns containing 'time' or 'stamp' case-insensitively.
    """
    for col in df.columns:
        if isinstance(col, str) and (
            "time" in col.lower() or "stamp" in col.lower()
        ):
            return col
    raise ValueError(
        f"Could not detect timestamp column. Columns: {df.columns.tolist()}"
    )

#  parse timestamp

def _parse_file_timestamps(df: pd.DataFrame, ts_col: str) -> tuple:
    """
    Parse timestamp column if it is stored as string.
    Returns (df with parsed timestamps, warning string or None)
    """
    warning = None

    # already datetime - openpyxl parsed it correctly
    if pd.api.types.is_datetime64_any_dtype(df[ts_col]):
        return df, warning
    if hasattr(df[ts_col].iloc[0], 'year'):
        try:
            df[ts_col] = pd.to_datetime(df[ts_col], errors="coerce")
            return df, warning
        except Exception:
            pass  # fall through to format loop

    # string timestamps - try formats in order
    # MM/DD/YYYY must come before DD/MM/YYYY
    # because your data is always American format
    formats = [
        "%m/%d/%Y %I:%M:%S %p",   # 5/1/2026 12:00:00 AM
        "%m/%d/%Y %I:%M %p",      # 5/1/2026 12:00 AM
        "%m/%d/%Y %H:%M:%S",      # 5/1/2026 00:00:00
        "%m/%d/%Y %H:%M",         # 5/1/2026 00:00
        "%Y-%m-%d %H:%M:%S",      # 2026-05-01 00:00:00
        "%Y-%m-%d %H:%M",         # 2026-05-01 00:00
    ]

    sample = str(df[ts_col].iloc[0])
    for fmt in formats:
        try:
            parsed = pd.to_datetime(df[ts_col], format=fmt, errors="raise")
            df = df.copy()
            df[ts_col] = parsed

            # sanity check - if month > 12 after parsing, something is wrong
            if df[ts_col].dt.month.max() > 12:
                warning = (
                    f"Timestamp parsing produced invalid months. "
                    f"Sample raw value: {sample}"
                )
            return df, warning
        except Exception:
            continue

    # last resort
    try:
        df = df.copy()
        df[ts_col] = pd.to_datetime(df[ts_col], infer_datetime_format=True)
        warning = (
            f"Timestamp format was inferred automatically. "
            f"Please verify dates are correct. Sample: {sample}"
        )
        return df, warning
    except Exception:
        raise ValueError(
            f"Could not parse timestamp column. "
            f"Sample value: {sample}"
        )




# -- Filter --------------------------------------------------------------------

def filter_data(df: pd.DataFrame) -> pd.DataFrame:
    """
    Apply time window and GII filter.
    Expects normalised column names (timestamp, gii, active_power_kw).
    """
    df   = df.copy()
    hour = df["timestamp"].dt.hour
    df   = df[(hour >= TIME_START) & (hour < TIME_END)]
    df   = df[df["gii"] > GII_FILTER]
    df   = df.dropna(subset=["active_power_kw"])
    return df.reset_index(drop=True)


# -- Time features ------------------------------------------------------------─

def add_time_features(df: pd.DataFrame) -> pd.DataFrame:
    """
    Add cyclical time encodings.
    XGBoost needs these explicitly - no inductive bias for time.
    """
    df  = df.copy()
    dt  = pd.to_datetime(df["timestamp"])
    hf  = dt.dt.hour + dt.dt.minute / 60.0

    df["hour_sin"]   = np.sin(2 * np.pi * hf / 24)
    df["hour_cos"]   = np.cos(2 * np.pi * hf / 24)
    df["month_sin"]  = np.sin(2 * np.pi * dt.dt.month / 12)
    df["month_cos"]  = np.cos(2 * np.pi * dt.dt.month / 12)
    df["doy_sin"]    = np.sin(2 * np.pi * dt.dt.dayofyear / 365)
    df["doy_cos"]    = np.cos(2 * np.pi * dt.dt.dayofyear / 365)
    df["hour_raw"]   = dt.dt.hour
    df["minute_raw"] = dt.dt.minute

    return df


# -- Feature columns ----------------------------------------------------------─

def get_feature_cols(df: pd.DataFrame) -> list:
    return [c for c in df.columns if c not in EXCLUDE_FROM_FEATURES]


# -- Training window ----------------------------------------------------------─

def apply_training_window(df: pd.DataFrame) -> pd.DataFrame:
    """
    Expanding window for first 12 months.
    Rolling 120-day window after 12 months.
    """
    df               = df.sort_values("timestamp").reset_index(drop=True)
    earliest         = df["timestamp"].min()
    latest           = df["timestamp"].max()
    months_available = (latest - earliest).days / 30.0

    if months_available >= ROLLING_WINDOW_TRIGGER_MONTHS:
        cutoff = latest - pd.Timedelta(days=ROLLING_WINDOW_DAYS)
        df     = df[df["timestamp"] >= cutoff].reset_index(drop=True)
        log.info(
            f"Rolling window active - using last {ROLLING_WINDOW_DAYS} days "
            f"({cutoff.date()} to {latest.date()})"
        )
    else:
        log.info(
            f"Expanding window - {months_available:.1f} months available, "
            f"using all data"
        )

    return df


# -- Blocked split ------------------------------------------------------------─

def split_blocked(df: pd.DataFrame) -> tuple:
    """
    Blocked temporal split.
    Train = all history except last (VAL_DAYS + TEST_DAYS)
    Val   = next VAL_DAYS
    Test  = last TEST_DAYS
    GAP_ROWS dropped at each boundary to prevent leakage.
    """
    df           = df.sort_values("timestamp").reset_index(drop=True)
    dates        = df["timestamp"].dt.date
    unique_dates = sorted(dates.unique())
    n            = len(unique_dates)

    if n < VAL_DAYS + TEST_DAYS + 7:
        raise ValueError(
            f"Not enough days for blocked split. "
            f"Need at least {VAL_DAYS + TEST_DAYS + 7}, got {n}."
        )

    test_dates  = set(unique_dates[-TEST_DAYS:])
    val_dates   = set(unique_dates[-(TEST_DAYS + VAL_DAYS):-TEST_DAYS])
    train_dates = set(unique_dates[:-(TEST_DAYS + VAL_DAYS)])

    train = df[dates.isin(train_dates)].reset_index(drop=True)
    val   = df[dates.isin(val_dates)].reset_index(drop=True)
    test  = df[dates.isin(test_dates)].reset_index(drop=True)

    train = train.iloc[:-GAP_ROWS] if len(train) > GAP_ROWS else train
    val   = val.iloc[GAP_ROWS:]    if len(val)   > GAP_ROWS else val
    test  = test.iloc[GAP_ROWS:]   if len(test)  > GAP_ROWS else test

    return train, val, test


# -- Full prepare pipeline (training + retrain) --------------------------------

def prepare(df: pd.DataFrame) -> tuple:
    """
    Full preparation pipeline for training and retraining:
    filter -> time features -> training window -> split

    Returns train, val, test DataFrames and feature column list.
    """
    df           = filter_data(df)
    df           = add_time_features(df)
    df           = apply_training_window(df)
    feature_cols = get_feature_cols(df)
    train, val, test = split_blocked(df)

    return train, val, test, feature_cols


# -- Prepare for inference (no split, no training window) --------------------─

def prepare_inference(df: pd.DataFrame, feature_cols: list) -> pd.DataFrame:
    """
    Filter and add time features for inference data.
    Uses saved feature_cols from training to ensure column consistency.
    Does not apply training window - use all uploaded inference data.
    """
    df = filter_data(df)
    df = add_time_features(df)

    missing = [c for c in feature_cols if c not in df.columns]
    if missing:
        raise ValueError(
            f"Inference data is missing columns that were present at training: "
            f"{missing}. Ensure your data files match the training format."
        )

    return df